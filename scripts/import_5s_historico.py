"""
Importa las 47 auditorías 5S históricas del Google Form (Excel) a Supabase.

Pre-requisito: aplicar APLICAR_5S_AUDITORES_EXTERNOS.sql en Supabase SQL Editor.

Estrategia: histórico simplificado.
  - Estado = 'completada'
  - nota_total y notas_por_s pre-calculados desde el form
  - NO se insertan filas en s5_auditoria_items
  - Auditor = registro nuevo en s5_auditores ('Cyro Michajlow')
"""
import json
import os
import re
import ssl
import sys
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path

import openpyxl

XLSX = Path("/root/Mercosur distribuciones/Entrega/AUDITORÍA 5 S (Respuestas).xlsx")
EVIDENCIAS_DIR = Path("/tmp/5s_evidencia")
SUPABASE_URL = os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
if not SUPABASE_URL or not SERVICE_ROLE_KEY:
    sys.exit("Faltan envs SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY")
AUDITOR_NOMBRE = "Cyro Michajlow"
MAX_PUNTAJE_FLOTA = 3
STORAGE_BUCKET = "s5-auditorias"

# Mapping de las 9 preguntas del form a categorías 5S
# (cols 41..49 del Excel, base 1)
PREGUNTAS_CATEGORIA = {
    41: "limpieza",        # plataforma libre de suciedad
    42: "limpieza",        # nivel limpieza exterior
    43: "organizacion",    # cabina libre de materiales innecesarios
    44: "orden",           # materiales ordenados en lugares establecidos
    45: "limpieza",        # superficies internas libres de suciedad
    46: "limpieza",        # piso libre de suciedad
    47: "limpieza",        # nivel limpieza interior
    48: "orden",           # elementos de limpieza suficientes (kit)
    49: "limpieza",        # buen estado general
}
COL_MARCA = 1
COL_CD = 2
COL_DOMINIO = 3
COL_CHOFER = 4
COL_AYUDANTE = 5
COL_EVIDENCIA = 51

ctx = ssl.create_default_context()


def sb(method, path, body=None, params=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    headers = {
        "apikey": SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, method=method, data=data, headers=headers)
    try:
        with urllib.request.urlopen(req, context=ctx) as r:
            text = r.read().decode()
            return json.loads(text) if text else None
    except urllib.error.HTTPError as e:
        msg = e.read().decode()
        raise RuntimeError(f"{method} {url} -> {e.code}: {msg}") from None


def storage_upload(bucket: str, path: str, body: bytes, mime: str = "image/jpeg") -> str:
    url = f"{SUPABASE_URL}/storage/v1/object/{bucket}/{path}"
    headers = {
        "apikey": SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
        "Content-Type": mime,
        "x-upsert": "true",
    }
    req = urllib.request.Request(url, method="POST", data=body, headers=headers)
    try:
        with urllib.request.urlopen(req, context=ctx) as r:
            r.read()
        return path
    except urllib.error.HTTPError as e:
        msg = e.read().decode()
        raise RuntimeError(f"upload {bucket}/{path} -> {e.code}: {msg}") from None


def storage_ensure_bucket(bucket: str):
    url = f"{SUPABASE_URL}/storage/v1/bucket"
    headers = {
        "apikey": SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
    }
    body = json.dumps({"id": bucket, "name": bucket, "public": False}).encode()
    req = urllib.request.Request(url, method="POST", data=body, headers=headers)
    try:
        with urllib.request.urlopen(req, context=ctx) as r:
            r.read()
        print(f"✓ Bucket creado: {bucket}")
    except urllib.error.HTTPError as e:
        if e.code in (409, 400):
            return  # ya existe
        msg = e.read().decode()
        raise RuntimeError(f"ensure bucket {bucket} -> {e.code}: {msg}") from None


def ensure_schema_ready():
    try:
        sb("GET", "s5_auditores", params={"select": "id", "limit": "1"})
    except Exception as e:
        print("✗ La migration parece no estar aplicada:", e)
        print("  Aplicá APLICAR_5S_AUDITORES_EXTERNOS.sql en Supabase SQL Editor.")
        sys.exit(1)
    cols = sb("GET", "s5_auditorias", params={"select": "auditor_externo_id", "limit": "1"})
    if cols is None:
        print("✗ Falta auditor_externo_id en s5_auditorias.")
        sys.exit(1)
    print("✓ Schema listo")


def upsert_auditor():
    existing = sb("GET", "s5_auditores", params={"select": "id,nombre", "nombre": f"eq.{AUDITOR_NOMBRE}"})
    if existing:
        print(f"✓ Auditor ya existe: {AUDITOR_NOMBRE} ({existing[0]['id']})")
        return existing[0]["id"]
    created = sb("POST", "s5_auditores", body={"nombre": AUDITOR_NOMBRE, "activo": True})
    print(f"✓ Auditor creado: {AUDITOR_NOMBRE} ({created[0]['id']})")
    return created[0]["id"]


def get_vehiculos_map():
    rows = sb("GET", "catalogo_vehiculos", params={"select": "id,dominio", "limit": "200"})
    return {r["dominio"]: r["id"] for r in rows}


def first_day(d: date) -> str:
    return date(d.year, d.month, 1).isoformat()


def calcular_notas(puntajes_por_pregunta: dict) -> tuple[float, dict]:
    """
    puntajes_por_pregunta: {col: int}  (col 41..49, valores 1 o 3)
    Retorna (nota_total %, notas_por_s {categoria: %})
    """
    acum = {}
    total_sum = 0
    total_n = 0
    for col, valor in puntajes_por_pregunta.items():
        if valor is None:
            continue
        cat = PREGUNTAS_CATEGORIA[col]
        pct = (valor / MAX_PUNTAJE_FLOTA) * 100
        acum.setdefault(cat, []).append(pct)
        total_sum += pct
        total_n += 1
    notas_por_s = {cat: round(sum(v) / len(v), 2) for cat, v in acum.items()}
    nota_total = round(total_sum / total_n, 2) if total_n else 0.0
    return nota_total, notas_por_s


def extract_drive_id(url: str) -> str | None:
    m = re.search(r"id=([\w\-]+)", url or "")
    return m.group(1) if m else None


def main():
    ensure_schema_ready()
    auditor_externo_id = upsert_auditor()
    vehiculos = get_vehiculos_map()
    storage_ensure_bucket(STORAGE_BUCKET)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Respuestas de formulario 1"]

    inserts = []
    evidencias = []
    skipped = []

    for r in range(3, ws.max_row + 1):
        marca = ws.cell(row=r, column=COL_MARCA).value
        if not marca:
            continue
        cd = ws.cell(row=r, column=COL_CD).value
        dominio = ws.cell(row=r, column=COL_DOMINIO).value
        chofer = ws.cell(row=r, column=COL_CHOFER).value
        ayudante = ws.cell(row=r, column=COL_AYUDANTE).value
        evidencia = ws.cell(row=r, column=COL_EVIDENCIA).value

        if dominio not in vehiculos:
            skipped.append((marca, dominio, "vehiculo no encontrado"))
            continue

        puntajes = {col: ws.cell(row=r, column=col).value for col in PREGUNTAS_CATEGORIA}
        if any(p is None for p in puntajes.values()):
            skipped.append((marca, dominio, "puntajes incompletos"))
            continue

        nota_total, notas_por_s = calcular_notas(puntajes)
        fecha_iso = marca.date().isoformat()

        observ = f"Importada desde Google Forms (Excel histórico). CD: {cd}."

        evidencia_storage_path = None
        if evidencia:
            drive_id = extract_drive_id(evidencia)
            local = EVIDENCIAS_DIR / f"{drive_id}.jpg"
            if drive_id and local.exists():
                evidencia_storage_path = (
                    f"historico/{marca.year}/{marca.month:02d}/{dominio}_{drive_id}.jpg"
                )
                evidencias.append((local, evidencia_storage_path, dominio, fecha_iso))
            else:
                observ += f" Evidencia original (no descargada): {evidencia}"

        inserts.append({
            "tipo": "flota",
            "periodo": first_day(marca.date()),
            "fecha": fecha_iso,
            "auditor_id": None,
            "auditor_externo_id": auditor_externo_id,
            "vehiculo_id": vehiculos[dominio],
            "chofer_nombre": (chofer or "").strip() or None,
            "ayudante_1": (ayudante or "").strip() or None,
            "ayudante_2": None,
            "sector_numero": None,
            "estado": "completada",
            "nota_total": nota_total,
            "notas_por_s": notas_por_s,
            "observaciones_generales": observ,
            "evidencia_storage_path": evidencia_storage_path,
            "created_at": marca.isoformat(),
        })

    print(f"\nResumen previo a insertar:")
    print(f"  Auditorías a insertar: {len(inserts)}")
    print(f"  Evidencias a subir a Storage: {len(evidencias)}")
    print(f"  Saltadas: {len(skipped)}")
    for s in skipped:
        print(f"    {s}")

    if "--dry-run" in sys.argv:
        print("\n[dry-run] no se insertó nada.")
        if inserts:
            print("Primer registro de muestra:")
            print(json.dumps(inserts[0], indent=2, default=str))
        return

    # Idempotencia: borrar previas del mismo auditor externo antes de re-cargar
    if "--reset" in sys.argv:
        print(f"\n[reset] borrando auditorías previas con auditor_externo_id={auditor_externo_id}")
        sb("DELETE", "s5_auditorias", params={"auditor_externo_id": f"eq.{auditor_externo_id}"})

    # 1) Subir evidencias a Storage
    print()
    for local, path, dom, f in evidencias:
        body = local.read_bytes()
        storage_upload(STORAGE_BUCKET, path, body, mime="image/jpeg")
        print(f"  ↑ evidencia subida {dom} {f}  ({len(body)} bytes) -> {path}")

    # 2) Insert auditorías por lotes
    BATCH = 20
    inserted = 0
    for i in range(0, len(inserts), BATCH):
        chunk = inserts[i:i+BATCH]
        result = sb("POST", "s5_auditorias", body=chunk)
        inserted += len(result)
        print(f"  insertados {inserted}/{len(inserts)}")

    print(f"\n✓ Carga completada: {inserted} auditorías 5S")
    print(f"  Evidencias subidas: {len(evidencias)}")


if __name__ == "__main__":
    main()
